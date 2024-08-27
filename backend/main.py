from flask import request, jsonify
from config import app, db
from models import UserData, TicketSale
import requests
import json
from sqlalchemy import inspect, func
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

class TicketNumber():
    def __init__(self, ticket_number):
        self.ticket_number = ticket_number

class TempTicketNumber():
    def __init__(self, ticket_number):
        self.ticket_number = ticket_number

count = TicketNumber(1)
temp = TempTicketNumber(None)

month = f'{datetime.now().month:02}'
date = f'{datetime.now().day:02}'
json_name = 'sample.json'
ticket_sale = None

def get_ticket():
    if temp.ticket_number is not None:
        temp_value = temp.ticket_number
        temp.ticket_number = None
        count.ticket_number -= 1
        return temp_value
    else:
        return count.ticket_number
    
def add_new_user_Excel(userID, name, status):
    try:
        workbook = load_workbook('전체이용자(편집).xlsx')
        sheet = workbook['회원']

        userID = int(userID)
        # Set Retrieved values to excel
        last_row = sheet.max_row + 1
        sheet[f'A{last_row}'].value = userID
        sheet[f'B{last_row}'].value = name
        sheet[f'C{last_row}'].value = status

        # Change font
        sheet[f'A{last_row}'].font = Font(name='굴림체', size=9, bold=False)
        sheet[f'B{last_row}'].font = Font(name='굴림체', size=9, bold=False)
        sheet[f'C{last_row}'].font = Font(name='굴림체', size=9, bold=False)

        # Align texts to center
        sheet[f'A{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'B{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'C{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
        workbook.save('전체이용자(편집).xlsx')
        return None
    except:
        return "Error Adding User"

def save_Excel():
    global month, date, ticket_sale
    try:
        # Fetch all records from the TicketSale table
        ticket_sales = ticket_sale.query.all()

        # Convert the list of TicketSale objects to a list of dictionaries
        ticket_sales_list = [ticket.to_json() for ticket in ticket_sales]

        

        file_name = f'전체이용자(편집)_{month}_{date}.xlsx'
        # Define the font properties
        font = Font(name='굴림체', size=9, bold=False)
        nm_font = Font(name='맑은 고딕', size=11, bold=False)

        workbook = load_workbook('전체이용자(편집).xlsx')
        sheet = workbook['회원']
        nonSheet = workbook['비회원']
        nm_row = 2

        for data in ticket_sales_list:
            userID = data['userID']
            name = data['name']
            status = data['status']
            price = int(data['price'])
            ticketNumber = int(data['ticketNumber'])

            if str(userID) == '비회원':
                nonSheet[f'A{nm_row}'].value = name
                nonSheet[f'B{nm_row}'].value = status
                nonSheet[f'C{nm_row}'].value = price
                nonSheet[f'D{nm_row}'].value = ticketNumber
                nonSheet[f'A{nm_row}'].font = nm_font
                nonSheet[f'B{nm_row}'].font = nm_font
                nonSheet[f'C{nm_row}'].font = nm_font
                nonSheet[f'D{nm_row}'].font = nm_font
                nm_row += 1
                continue
            for cell in sheet['A']:
                if str(cell.value) == str(userID):
                    sheet[f'D{cell.row}'].value = price
                    sheet[f'E{cell.row}'].value = ticketNumber
                    sheet[f'D{cell.row}'].font = font
                    sheet[f'E{cell.row}'].font = font
        
        # Save workbook
        workbook.save(f'{file_name}')
        return None
    except:
        return "Error Saving File"

@app.route('/sell_ticket', methods=['POST'])
def sell_ticket():
    global ticket_sale
    user_id = request.json.get("userID")
    name = request.json.get("name")
    status = request.json.get("status")
    price = request.json.get("price")
    ticket_number = request.json.get("ticketNumber")

    if user_id is None or name is None or price is None:
        return jsonify({"message": "Some entry is missing in API request"}), 400
    
    sale_data = ticket_sale(user_id=user_id, name=name, status=status, price=price, ticket_number=ticket_number)
    try:
        db.session.add(sale_data)
        db.session.commit()
        count.ticket_number += 1
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    
    return jsonify({"message": sale_data}), 201

@app.route('/get_ticket_number', methods=['GET'])
def get_ticket_number():
    if count.ticket_number is None:
        return jsonify({"message": "Error retrieving Ticket Number"}), 401
    return jsonify({"message": count.ticket_number}), 201

@app.route('/reset_ticket_number/<int:ticket_number>', methods=['POST'])
def reset_ticket_number(ticket_number):
    count.ticket_number = ticket_number
    
    if count.ticket_number != None:
        return jsonify({"message": count.ticket_number}), 201
    else:
        return jsonify({"message": "Error reseting ticketNumber"}), 401

@app.route('/non_member/<string:name>/<string:purpose_of_visit>', methods=['POST'])
def non_member(name, purpose_of_visit):
    if name is None or purpose_of_visit is None:
        return jsonify({"message": "Some entry is Missing"}), 400
    
    sale_data_dict = {
        "userID": "비회원",
        "name": name,
        "status": purpose_of_visit,
        "price": 3500,
        "ticketNumber": get_ticket()
    }

    data = add_json(sale_data_dict)

    response = {
        "table": data,
        "sales": sale_data_dict,
    }
    
    return jsonify(response), 201

def add_json(dict):
    global json_name
    with open(json_name, 'r', encoding='utf-8') as file:
        data = json.load(file)
    
    data.append(dict)

    with open(json_name, 'w', encoding='utf-8') as file:
            json.dump(data, file, ensure_ascii=False)

    return data

def deleteJSON(user_id):
    global json_name
    with open(json_name, 'r', encoding='utf-8') as file:
            user_data = json.load(file)

    # Creates a new list that only includes data that does not have matching userID    
    user_data = [data for data in user_data if data['userID'] != user_id]

    # Write the updated data back to the file
    with open(json_name, 'w', encoding='utf-8') as file:
        json.dump(user_data, file, indent=4, ensure_ascii=False)
    
    return user_data

@app.route('/sell_ticket_json/<user_id>', methods=['GET'])
def sell_ticket_json(user_id):
    user = UserData.query.filter_by(user_id=user_id).first()
    if user:
        sale_data = {
            "userID": user.user_id,
            "name": user.name,
            "status": user.status,
            "price": user.price,
            "ticketNumber": get_ticket()
        }
        data = add_json(sale_data)
        response = {
            "sales": sale_data,
            "table": data
        }
        return jsonify(response), 201
    else:
        return jsonify({"message": "User not found"}), 404

@app.route('/add_sale_db', methods=['POST'])
def add_sale_db():
    global ticket_sale
    user_id = request.json.get("userID")
    name = request.json.get("name")
    status = request.json.get("status")
    price = request.json.get("price")
    ticket_number = request.json.get("ticketNumber")

    if user_id is None or name is None or price is None:
        return jsonify({"message": "Some entry is missing in API request"}), 400
    
    sale_data = ticket_sale(user_id=user_id, name=name, status=status, price=price, ticket_number=ticket_number)
    try:
        db.session.add(sale_data)
        db.session.commit()
        
        count.ticket_number += 1
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    
    return jsonify({"message": "Data Added"}), 201

@app.route('/add_user', methods=['POST'])
def add_user():
    user_id = request.json.get("userID")
    name = request.json.get("name")
    status = request.json.get("status")
    price = request.json.get("price")

    if user_id is None or name is None or status is None or price is None:
        return jsonify({"message": "Some entry is missing in API request"}), 400
    
    new_user = UserData(user_id=user_id, name=name, status= status, price=price)
    try:
        db.session.add(new_user)
        db.session.commit()
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    return jsonify({"message": "User Added"}), 201

@app.route('/refund_ticket/<int:ticket_number>', methods=['DELETE'])
def refund_ticket(ticket_number):
    global ticket_sale
    user = ticket_sale.query.filter_by(ticket_number=int(ticket_number)).first()

    if user == None:
        return jsonify({"message": "User not found"}), 404
    
    db.session.delete(user)
    db.session.commit()

    temp.ticket_number = ticket_number
    
    data = deleteJSON(user.user_id)

    return jsonify(data), 201

@app.route('/add_new_user/<user_id>/<name>/<status>/', methods=['POST'])
def add_new_user(user_id, name, status):
    status_to_price = {
        '기초생활수급권자': 0,
        '차상위(저소득)': 0,
        '기타': 1750,
        '국가유공자': 1750,
        '일반': 3500
    }

    price = status_to_price.get(status, 9999)

    new_user = {
        "userID": user_id,
        "name": name,
        "status": status,
        "price": price
    }
    response = requests.post('http://localhost:5000/add_user', json=new_user)

    # Check the response status code and content
    if response.status_code == 201:
        print("Success:", response.json())
    elif response.status_code == 400:
        print("Failed:", response.status_code, response.json())

    result = add_new_user_Excel(user_id, name, status)

    if result != None:
        jsonify({"message": "Error Adding New User"}), 401

    sale_data = {
            "userID": user_id,
            "name": name,
            "status": status,
            "price": price,
            "ticketNumber": get_ticket()
        }
    data = add_json(sale_data)
    sale_response = requests.post('http://localhost:5000/add_sale_db', json=sale_data)

    # Check the response status code and content
    if sale_response.status_code == 201:
        print("Success:", response.json())
    elif sale_response.status_code == 400:
        print("Failed:", response.status_code, response.json())

    return jsonify(data), 201

@app.route('/save_file', methods=['POST'])
def save_file():
    result = save_Excel()

    if result != None:
        return jsonify({"message": result}), 401
    else:
        return jsonify({"message": "Saved File"}), 201

def startup_task():
    global month, date, ticket_sale
    table_name = f'ticketsale_{month}_{date}'
    ticket_sale = type(table_name, (TicketSale,), {'__tablename__': table_name})

    insepctor = inspect(db.engine)
    if table_name not in insepctor.get_table_names():
        db.create_all()

    max = db.session.query(func.max(ticket_sale.ticket_number)).scalar()
    if max != None:
        count.ticket_number = max + 1

if __name__ == "__main__":
    with app.app_context():
        startup_task()
    app.run(debug=True)